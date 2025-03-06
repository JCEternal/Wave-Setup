# Wave Report 1.0 3/5/25
# Created by JC
# Needs requests & openpyxl & pillow installed, pip install requests openpyxl pillow

# This script runs against Wave and gathers camera name, vendor, model, IP, MAC, firmware and a thumbnail of the image
# and saves it into a default location of c:\reports along with a folder of all the images.

# Optimized to run via windows command prompt

# Script running video for first timers - https://youtu.be/W1CuoEo-Jtw

import requests
import sys
import re
import os
import ipaddress
import pathlib
import platform
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image as OpenPyxlImage

import requests
import sys
import re
import os
import ipaddress
import pathlib
import platform
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image as OpenPyxlImage


def clear_console():
    if platform.system() == "Windows":
        os.system("cls")
    else:
        os.system("clear")


def sanitize_filename(name):
    # Remove characters not allowed in filenames (especially for Windows)
    return re.sub(r'[\\/*?:"<>|]', "", name)


def validate_directory_input(prompt, default):
    """
    Validate that the given file path is acceptable for Windows.
    It checks that the path parts (except a drive letter) do not contain invalid characters.
    """
    # Define invalid characters for directory names (drive letter is allowed colon)
    invalid_chars = '<>:"|?*'
    while True:
        path_input = input(prompt) or default
        try:
            p = pathlib.Path(path_input)
            # Resolve the path (without requiring it to exist)
            resolved = p.resolve(strict=False)
            # Check each part of the path
            for part in p.parts:
                # Allow drive letters like "C:\" or "C:" (regex adjusted to accept an optional backslash)
                if re.match(r'^[A-Za-z]:\\?$', part):
                    continue
                for ch in invalid_chars:
                    if ch in part:
                        raise ValueError(f"Path part '{part}' contains invalid character: {ch}")
            return str(resolved)
        except Exception as e:
            print("Invalid file path:", e)


def authenticate_wave(base_url, username, password, headers):
    auth_url = f"{base_url}/login/sessions"
    auth_payload = {"username": username, "password": password}
    try:
        response = requests.post(auth_url, json=auth_payload, headers=headers, verify=False)
        if response.status_code in [200, 201]:
            token = response.json().get("token")
            headers["Authorization"] = f"Bearer {token}"
            print("Authentication successful.")
            return True
        else:
            print("Authentication failed. Exiting...")
            return False
    except Exception as e:
        print("Error during authentication:", e)
        return False


def get_devices(base_url, headers):
    devices_url = f"{base_url}/devices"
    try:
        response = requests.get(devices_url, headers=headers, verify=False)
        if response.status_code == 200:
            devices = response.json()
            if isinstance(devices, list):
                return devices
            else:
                print("Unexpected format for devices data.")
                return []
        else:
            print("Failed to retrieve devices. Status code:", response.status_code)
            return []
    except Exception as e:
        print("Error fetching devices:", e)
        return []


def fetch_device_image(base_url, device, headers, images_dir, file_name):
    # Retrieve image using the unique device id, but save file using camera name.
    image_url = f"{base_url}/devices/{device.get('id')}/image?timestampMs=-1&rotation=10&size=240x-1&streamSelectionMode=forcedSecondary"
    try:
        response = requests.get(image_url, headers=headers, verify=False)
        if response.status_code == 200:
            image_path = os.path.join(images_dir, file_name)
            with open(image_path, "wb") as img_file:
                img_file.write(response.content)
            if os.path.exists(image_path):
                return image_path
    except Exception as e:
        print(f"Error fetching image for device {device.get('id')}: {e}")
    return None


def export_to_excel(reports_dir, device_data, device_ids, device_images):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cameras"

    # Column order: Camera Name, Vendor, Model, IP Address, MAC Address, Firmware, Image
    headers_list = ["Camera Name", "Vendor", "Model", "IP Address", "MAC Address", "Firmware", "Image"]
    ws.append(headers_list)

    max_image_width = 0  # Track maximum image width for column G

    row = 2
    for i, data in enumerate(device_data):
        ws.append(data)
        # Get the corresponding device id (not exported in XLSX)
        device_id = device_ids[i]
        if device_id in device_images and device_images[device_id] and os.path.exists(device_images[device_id]):
            try:
                img = OpenPyxlImage(device_images[device_id])
                img.anchor = f"G{row}"  # Column G for the image
                ws.add_image(img, f"G{row}")
                # Set row height based on image height (approximate conversion: 1 point ~ 0.75 pixels)
                ws.row_dimensions[row].height = img.height * 0.75
                if img.width > max_image_width:
                    max_image_width = img.width
            except Exception as e:
                print(f"Error inserting image for device {device_id}: {e}")
                ws.row_dimensions[row].height = 80
        else:
            ws.row_dimensions[row].height = 80

        # Vertically center all cells in this row
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).alignment = Alignment(vertical='center')
        row += 1

    # Auto adjust all columns except the image column (G)
    for col in ws.columns:
        col_letter = col[0].column_letter
        if col_letter == 'G':
            continue
        max_length = max((len(str(cell.value)) if cell.value is not None else 0 for cell in col), default=0)
        ws.column_dimensions[col_letter].width = max_length + 2

    # Adjust column G width based on the maximum image width (approximation)
    if max_image_width:
        ws.column_dimensions['G'].width = (max_image_width - 5) / 7
    else:
        ws.column_dimensions['G'].width = 15

    report_file = os.path.join(reports_dir, "camera_report.xlsx")
    try:
        wb.save(report_file)
        print("Data export completed successfully.")
    except Exception as e:
        print("Error saving Excel file:", e)


def main():
    clear_console()
    # Require a valid IP address (no default)
    while True:
        server_ip = input("Enter Wave VMS Server IP: ").strip()
        try:
            ipaddress.ip_address(server_ip)
            break
        except ValueError:
            print("Invalid IP address. Please enter a valid IP address.")

    server_port = input("Enter Wave VMS Server Port (default: 7001): ") or "7001"
    server_username = input("Enter Wave VMS Username (default: admin): ") or "admin"

    # Require a non-empty password (no default)
    while True:
        server_password = input("Enter Wave VMS Password: ").strip()
        if server_password:
            break
        else:
            print("Server password cannot be empty. Please enter a valid password.")

    confirm_password = input("Confirm Wave VMS Password? (y/n): ").strip().lower()
    if confirm_password not in ['y', 'yes']:
        print("Password confirmation failed. Exiting...")
        sys.exit(1)

    base_url = f"https://{server_ip}:{server_port}/rest/v3"
    headers = {"Content-Type": "application/json"}
    requests.packages.urllib3.disable_warnings()

    if not authenticate_wave(base_url, server_username, server_password, headers):
        sys.exit(1)

    # Validate and create reports directory if it doesn't exist
    reports_dir = validate_directory_input("Enter location to store reports (default: C:\\reports): ", "C:\\reports")
    os.makedirs(reports_dir, exist_ok=True)
    images_dir = os.path.join(reports_dir, "images")
    os.makedirs(images_dir, exist_ok=True)

    devices = get_devices(base_url, headers)
    if not devices:
        sys.exit(1)

    device_data = []  # Each row: [camera_name, vendor, model, ip_address, mac, firmware]
    device_ids = []  # Corresponding device id for each row (for image lookup)
    device_images = {}

    print("\nDevice List:")
    print("{:<30} {:<20} {:<20} {:<30} {:<20} {:<15}".format(
        "Camera Name", "Vendor", "Model", "IP Address", "MAC Address", "Firmware"
    ))
    print("-" * 130)

    for device in devices:
        device_id = device.get("id", "Unknown")
        camera_name = device.get("name", "Unknown")
        mac = device.get("mac", "Unknown").lower().replace("-", ":")
        model = device.get("model", "Unknown")
        vendor = device.get("vendor", "Unknown")

        firmware = "Unknown"
        parameters = device.get("parameters", {})
        if isinstance(parameters, dict):
            firmware = parameters.get("firmware", "Unknown")

        ip_address = "Unknown"
        stream_urls = device.get("streamUrls", {})
        if isinstance(stream_urls, dict):
            for key, url in stream_urls.items():
                match = re.search(r'\b(?:[0-9]{1,3}\.){3}[0-9]{1,3}\b', url)
                if match:
                    ip_address = match.group()
                    break

        print("{:<30} {:<20} {:<20} {:<30} {:<20} {:<15}".format(
            camera_name, vendor, model, ip_address, mac, firmware
        ))

        # Save the image using the camera name (sanitized)
        file_name = f"{sanitize_filename(camera_name)}.jpg"
        image_path = fetch_device_image(base_url, device, headers, images_dir, file_name)
        device_images[device_id] = image_path

        # Append row data (without device id)
        device_data.append([camera_name, vendor, model, ip_address, mac, firmware])
        device_ids.append(device_id)

    export_to_excel(reports_dir, device_data, device_ids, device_images)

    # Pause at end so the Command Prompt window stays open
    if platform.system() == "Windows":
        os.system("pause")


if __name__ == "__main__":
    main()
