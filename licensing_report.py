# Wave Licensing Report 1.0 3/5/25
# Created by JC
# Needs requests & openpyxl & pandas installed, pip install requests openpyxl pandas

# This script authenticates to the wave server, gathers license data and exports into an xlsx file.

# Optimized to run via windows command prompt

# Script running video for first timers - https://youtu.be/W1CuoEo-Jtw

import os
import sys
import requests
import urllib3
import pandas as pd
import ipaddress
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Disable insecure request warnings (for development purposes only)
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


def authenticate_wave(base_url, username, password, headers):
    auth_url = f"{base_url}/rest/v3/login/sessions"
    auth_payload = {"username": username, "password": password}
    try:
        response = requests.post(auth_url, json=auth_payload, headers=headers, verify=False)
        if response.status_code in [200, 201]:
            token = response.json().get("token")
            headers["Authorization"] = f"Bearer {token}"
            print("Authentication successful.")
            return True
        else:
            print("Authentication failed. Exiting...")
            return False
    except Exception as e:
        print("Error during authentication:", e)
        return False


def parse_license_block(license_block):
    """
    Parses a multi-line licenseBlock string into a dictionary.
    """
    result = {}
    for line in license_block.split("\n"):
        if "=" in line:
            key, value = line.split("=", 1)
            result[key.strip()] = value.strip()
    return result


def get_licenses(base_url, headers):
    licenses_url = f"{base_url}/rest/v3/licenses"
    try:
        response = requests.get(licenses_url, headers=headers, verify=False)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        print("Error retrieving licenses:", e)
        return []

    try:
        licenses_data = response.json()
    except ValueError as e:
        print("Error parsing JSON response:", e)
        return []

    # Handle both single dict and list responses.
    if isinstance(licenses_data, dict):
        licenses_list = [licenses_data]
    else:
        licenses_list = licenses_data

    extracted_licenses = []
    for license_item in licenses_list:
        license_block = license_item.get("licenseBlock", "")
        parsed_block = parse_license_block(license_block)
        extracted_license = {
            "License": parsed_block.get("SERIAL", "N/A"),
            "Count": parsed_block.get("COUNT", "N/A"),
            "Hardware ID": parsed_block.get("HWID", "N/A")
        }
        extracted_licenses.append(extracted_license)
        print(f"License: {extracted_license['License']}")
        print(f"Count: {extracted_license['Count']}")
        print(f"Hardware ID: {extracted_license['Hardware ID']}")
        print("-" * 30)
    return extracted_licenses


def get_license_usage_summary_all(base_url, headers):
    """
    Retrieves the global license usage summary from the wildcard endpoint.
    Expected response sample:
    {
      "trial": {
        "available": 16,
        "inUse": 4,
        "total": 16
      }
    }
    """
    usage_url = f"{base_url}/rest/v3/licenses/*/summary"
    try:
        response = requests.get(usage_url, headers=headers, verify=False)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        print("Error retrieving license usage summary:", e)
        return []

    try:
        usage_data = response.json()
    except ValueError as e:
        print("Error parsing JSON for usage summary:", e)
        return []

    usage_list = []
    # Convert each key (license type) into a row.
    for license_type, details in usage_data.items():
        usage_list.append({
            "Type": license_type,
            "Available": details.get("available", "N/A"),
            "In Use": details.get("inUse", "N/A"),
            "Total": details.get("total", "N/A")
        })
    return usage_list


def auto_adjust_column_widths(writer):
    """
    Auto adjust column widths for each sheet in the Excel workbook.
    """
    for sheet_name, worksheet in writer.sheets.items():
        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[col_letter].width = adjusted_width


def export_to_excel(licenses, usage_summaries, file_path):
    """
    Exports license details and usage summary to an Excel file with two sheets.
    Centers the data in the 'Count' column of the Licenses sheet and centers all columns on the Usage sheet.
    """
    licenses_df = pd.DataFrame(licenses)
    usage_df = pd.DataFrame(usage_summaries)

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        licenses_df.to_excel(writer, sheet_name="Licenses", index=False)
        usage_df.to_excel(writer, sheet_name="Usage", index=False)

        # Auto adjust columns for both sheets.
        auto_adjust_column_widths(writer)

        # Center the data in the 'Count' column on the Licenses sheet.
        ws_licenses = writer.sheets["Licenses"]
        count_col_index = None
        # Find the column index for header "Count" (1-indexed).
        for idx, cell in enumerate(ws_licenses[1], start=1):
            if cell.value == "Count":
                count_col_index = idx
                break
        if count_col_index:
            for row in range(2, ws_licenses.max_row + 1):
                ws_licenses.cell(row=row, column=count_col_index).alignment = Alignment(horizontal="center")

        # Center all columns on the Usage sheet.
        ws_usage = writer.sheets["Usage"]
        for row in ws_usage.iter_rows(min_row=1, max_row=ws_usage.max_row, min_col=1, max_col=ws_usage.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

    print(f"License data and usage summary exported to {file_path}")


def main():
    # Clear the screen for a clean CMD interface.
    if os.name == "nt":
        os.system("cls")

    # Prompt for connection details and credentials.
    # Server IP (required, no default) with validation.
    while True:
        server_ip = input("Enter server IP: ").strip()
        if not server_ip:
            print("Server IP is required.")
            continue
        try:
            ipaddress.ip_address(server_ip)
            break
        except ValueError:
            print("Invalid IP address. Please enter a valid IP address.")

    port_input = input("Enter port [7001]: ").strip()
    username_input = input("Enter username [admin]: ").strip()

    # Password (required, no default) with verification
    while True:
        password = input("Enter password: ").strip()
        if not password:
            print("Password is required.")
            continue
        verify_choice = input("Is this password correct? (y/n): ").strip().lower()
        if verify_choice == "y":
            break
        else:
            print("Please re-enter your password.")

    file_location_input = input("Enter file location [C:\\reports]: ").strip()

    # Use default values if input is empty.
    port = port_input if port_input else "7001"
    username = username_input if username_input else "admin"
    file_location = file_location_input if file_location_input else r"C:\reports"

    # Create the file location directory if it doesn't exist.
    if not os.path.exists(file_location):
        try:
            os.makedirs(file_location)
            print(f"Created directory: {file_location}")
        except Exception as e:
            print("Error creating directory:", e)
            return

    # Construct the base URL (using HTTPS).
    base_url = f"https://{server_ip}:{port}"
    headers = {"Content-Type": "application/json"}

    # Authenticate using the provided function.
    if not authenticate_wave(base_url, username, password, headers):
        return

    # Retrieve license information.
    licenses = get_licenses(base_url, headers)

    # Retrieve global license usage summary.
    usage_summaries = get_license_usage_summary_all(base_url, headers)

    # Export both datasets to an Excel file.
    if licenses or usage_summaries:
        file_path = os.path.join(file_location, "licenses.xlsx")
        export_to_excel(licenses, usage_summaries, file_path)

    input("Press Enter to exit...")


if __name__ == "__main__":
    main()
